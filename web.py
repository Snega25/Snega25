# # import requests
# # from bs4 import BeautifulSoup
# # import openpyxl

# # # Open the input Excel sheet and select the relevant worksheet
# # workbook = openpyxl.load_workbook('college data.xlsx')
# # worksheet = workbook['data'] # Change this to the name of your worksheet

# # # Loop over each row in the worksheet and extract the data
# # for row in worksheet.iter_rows(min_row=2, values_only=True):
# #     url = row[0]

# #     # Make a GET request to the URL
# #     response = requests.get(url)

# #     # Parse the HTML content using BeautifulSoup
# #     soup = BeautifulSoup(response.content, 'html.parser')

# #     # Extract the contact details and email IDs using BeautifulSoup selectors
# #     contact_details = soup.select('selector for contact details')
# #     email_ids = soup.select('selector for email IDs')

# #     # Write the extracted data to the output worksheet
# #     row_data = [str(url), str(contact_details), str(email_ids)]
# #     worksheet.append(row_data)

# # # Save the output Excel sheet
# # workbook.save('output.xlsx')


# import pandas as pd

# # Read the Excel file
# df = pd.read_excel('college data.xlsx', usecols=['https://pec.paavai.edu.in/ 
# ',https://kahedu.edu.in/,https://www.srcas.ac.in/,https://www.psgcas.ac.in/])

# # Extract the URLs
# urls = df['URLs'].tolist()
# import requests
# from bs4 import BeautifulSoup

# for url in urls:
#     # Make a GET request to the URL
#     response = requests.get(url)

#     # Parse the HTML content using BeautifulSoup
#     soup = BeautifulSoup(response.content, 'html.parser')

#     # Extract the contact details and email IDs using BeautifulSoup selectors
#     contact_details = soup.select('selector for contact details')
#     email_ids = soup.select('selector for email IDs')

#     # Print the extracted data
#     print(contact_details)
#     print(email_ids)
# import csv

# # Open a CSV file to store the extracted data
# with open('college_contacts.csv', mode='w', newline='') as file:
#     writer = csv.writer(file)

#     # Write the header row
#     writer.writerow(['URL', 'Contact Details', 'Email IDs'])

#     # Loop over each URL and extract the data
#     for url in urls:
#         # Make a GET request to the URL
#         response = requests.get(url)

#         # Parse the HTML content using BeautifulSoup
#         soup = BeautifulSoup(response.content, 'html.parser')

#         # Extract the contact details and email IDs using BeautifulSoup selectors
#         contact_details = soup.select('selector for contact details')
#         email_ids = soup.select('selector for email IDs')

#         # Write the extracted data to the CSV file
#         writer.writerow([url, contact_details, email_ids])

import requests
from bs4 import BeautifulSoup
import openpyxl

# Open the input Excel sheet and select the relevant worksheet
workbook = openpyxl.load_workbook('college data.xlsx')
worksheet = workbook['Sheet1'] # Change this to the name of your worksheet

# Loop over each row in the worksheet and extract the data
for row in worksheet.iter_rows(min_row=2, values_only=True):
    url = row[0]

    # Make a GET request to the URL
    response = requests.get(url)

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Extract the contact details and email IDs using BeautifulSoup selectors
    contact_details = soup.select('selector for contact details')
    email_ids = soup.select('selector for email IDs')

    # Write the extracted data to the output worksheet
    row_data = [str(url), str(contact_details), str(email_ids)]
    worksheet.append(row_data)

# Save the output Excel sheet
workbook.save('output.xlsx')

